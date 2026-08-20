#!/usr/bin/env python3
"""EQUIPA · Genera el catálogo maestro en Excel con modelo de costos de importación.
Entradas: /tmp/equipa-skus.json (SKUs de la página) + /tmp/equipa-fob.json (FOB/CBM por SKU, de la investigación)
Salida:   ~/claude/equipa/investigacion/EQUIPA-CATALOGO-MAESTRO.xlsx  (correr recalc.py después)
"""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

skus = json.load(open("/tmp/equipa-skus.json"))
fob  = json.load(open("/tmp/equipa-fob.json"))
comp = json.load(open("/Users/andremacouzet/claude/equipa/investigacion/catalogo-referencia.json"))

AR   = lambda b=False, sz=10, col="000000": Font(name="Arial", size=sz, bold=b, color=col)
AZUL = Font(name="Arial", size=10, color="0000FF")
AMAR = PatternFill("solid", fgColor="FFFF00")
GRISF= PatternFill("solid", fgColor="121212")
HDR  = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BORDE= Border(bottom=Side(style="thin", color="DDDDDD"))
MXN  = '$#,##0'
MXN2 = '$#,##0.00'
USD  = '#,##0.00'
PCT  = '0.0%'

wb = Workbook()

# ═══ LEEME ═══
ws = wb.active; ws.title = "LEEME"
ws.column_dimensions["A"].width = 110
filas = [
 ("EQUIPA · CATÁLOGO MAESTRO — modelo de costos de importación", True),
 ("", False),
 ("Qué es: los 39 SKUs del catálogo con su costo real de importación (FOB → flete → arancel → aterrizado) y la utilidad por pieza a precio de mayoreo y menudeo.", False),
 ("", False),
 ("CÓMO USAR:", True),
 ("1. Celdas AZULES con fondo AMARILLO = se editan (parámetros, FOB, precios). Todo lo demás son fórmulas: NO tocar.", False),
 ("2. Cambia un parámetro en PARAMETROS (ej. arancel 25%→35%) y todo el catálogo se recalcula.", False),
 ("3. Columna 'Dato' en CATALOGO: [V] = FOB tomado de listado real verificable (URL en 'Fuente') · [E] = estimación de la investigación. ", False),
 ("   Los [E] se vuelven [V] cuando el proveedor responda el RFQ (ventas/rfq-proveedores.md) con cotización formal.", False),
 ("4. CONTENEDOR simula el primer 40'HQ: edita cantidades (azul) y revisa que el CBM no pase el límite.", False),
 ("5. COMPETENCIA: las 465 piezas de mobiliarioenmexico.com.mx por categoría, para validar precios de venta.", False),
 ("", False),
 ("Fuentes de parámetros: investigacion/sourcing-china.md y mercado-mexico.md (flete Efanda/Sino-Shipping ago-2026; arancel decreto DOF 29-dic-2025; DTA/agente Camtom y DTD Express). El IGI exacto por fracción lo confirma el agente aduanal.", False),
 ("Generado: 20-ago-2026 · se regenera con investigacion/armar_excel.py", False),
]
for i,(t,b) in enumerate(filas,1):
    ws.cell(i,1,t).font = AR(b, 12 if i==1 else 10)
    ws.cell(i,1).alignment = Alignment(wrap_text=True, vertical="top")

# ═══ PARAMETROS ═══
ws = wb.create_sheet("PARAMETROS")
ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 14; ws.column_dimensions["C"].width = 66
ws.cell(1,1,"PARÁMETROS DEL MODELO (editar en azul)").font = AR(True, 12)
params = [
 # (fila, etiqueta, valor, formato, nota)
 (3,"Tipo de cambio MXN/USD", 18.5, '0.00', "Ajustar al día de la operación"),
 (4,"Flete marítimo 40'HQ China→Manzanillo (USD)", 5000, USD, "Rango ago-2026: $3,500–5,500 (Efanda/Sino-Shipping). Cotizar en firme"),
 (5,"Seguro (% sobre FOB)", 0.003, PCT, "~0.3% típico"),
 (6,"IGI / arancel (% sobre CIF)", 0.30, PCT, "Decreto DOF 29-dic-2025: muebles cap. 94 quedaron en 25–35% según fracción. 30% = escenario medio. CONFIRMAR por fracción con agente aduanal"),
 (7,"DTA (% sobre CIF)", 0.008, PCT, "8 al millar, sin tope"),
 (8,"Agente aduanal + prevalidación + maniobras (USD/contenedor)", 900, USD, "MXN $3,500–8,000 + gastos; escenario medio"),
 (9,"Flete Manzanillo→Morelia contenedor (USD)", 1800, USD, "~MXN $25,000–40,000 [E]"),
 (10,"CBM útil por 40'HQ", 67, '0', "Nominal 76 CBM; 67 realista con estiba mixta"),
 (11,"IVA importación (acreditable)", 0.16, PCT, "Sobre CIF+IGI+DTA. Es flujo, no costo, si facturas — por eso NO entra al costo aterrizado"),
]
for f, lbl, val, fmt, nota in params:
    ws.cell(f,1,lbl).font = AR()
    c = ws.cell(f,2,val); c.font = AZUL; c.fill = AMAR; c.number_format = fmt
    ws.cell(f,3,nota).font = AR(sz=9, col="666666")
P_ = lambda fila: f"PARAMETROS!$B${fila}"

# ═══ CATALOGO ═══
ws = wb.create_sheet("CATALOGO")
cols = ["SKU","Producto","Categoría","FOB USD/ud","CBM/ud","Uds por 40'HQ","Flete USD/ud","Seguro USD","CIF USD",
        "IGI USD","DTA USD","Fijos USD/ud","Aterrizado USD","Aterrizado MXN","Precio mayoreo MXN","Precio menudeo MXN",
        "Utilidad may. $","Margen may. %","Utilidad men. $","Margen men. %","MOQ prov.","Dato","Fuente FOB (URL)"]
anchos = [9,30,12,11,8,12,11,10,10,10,9,11,12,13,15,15,13,11,13,11,9,6,60]
for j,(t,a) in enumerate(zip(cols,anchos),1):
    c = ws.cell(1,j,t); c.font = HDR; c.fill = GRISF; c.alignment = Alignment(wrap_text=True, vertical="center")
    ws.column_dimensions[get_column_letter(j)].width = a
ws.row_dimensions[1].height = 30
ws.freeze_panes = "D2"

for i, s in enumerate(skus, 2):
    d = fob.get(s["sku"], {})
    ws.cell(i,1,s["sku"]).font = AR(True)
    ws.cell(i,2,s["nombre"].replace(" "+s["sku"],"")).font = AR()
    ws.cell(i,3,s["cat"]).font = AR()
    c = ws.cell(i,4, d.get("fob")); c.font = AZUL; c.fill = AMAR; c.number_format = USD
    c = ws.cell(i,5, d.get("cbm")); c.font = AZUL; c.fill = AMAR; c.number_format = '0.000'
    ws.cell(i,6, f"=ROUNDDOWN({P_(10)}/E{i},0)").number_format = '#,##0'
    ws.cell(i,7, f"={P_(4)}/F{i}").number_format = USD
    ws.cell(i,8, f"=D{i}*{P_(5)}").number_format = USD
    ws.cell(i,9, f"=D{i}+G{i}+H{i}").number_format = USD
    ws.cell(i,10,f"=I{i}*{P_(6)}").number_format = USD
    ws.cell(i,11,f"=I{i}*{P_(7)}").number_format = USD
    ws.cell(i,12,f"=({P_(8)}+{P_(9)})/F{i}").number_format = USD
    ws.cell(i,13,f"=I{i}+J{i}+K{i}+L{i}").number_format = USD
    ws.cell(i,14,f"=M{i}*{P_(3)}").number_format = MXN2
    c = ws.cell(i,15, s["precio"]); c.font = AZUL; c.fill = AMAR; c.number_format = MXN
    c = ws.cell(i,16, s["antes"]);  c.font = AZUL; c.fill = AMAR; c.number_format = MXN
    ws.cell(i,17,f"=O{i}-N{i}").number_format = MXN
    ws.cell(i,18,f"=IF(O{i}=0,0,Q{i}/O{i})").number_format = PCT
    ws.cell(i,19,f"=P{i}-N{i}").number_format = MXN
    ws.cell(i,20,f"=IF(P{i}=0,0,S{i}/P{i})").number_format = PCT
    ws.cell(i,21, d.get("moq")).font = AR()
    ws.cell(i,22, d.get("dato","[E]")).font = AR()
    ws.cell(i,23, d.get("fuente","")).font = AR(sz=8, col="666666")
    for j in range(6,15):
        ws.cell(i,j).font = AR()
    for row in [17,19]: ws.cell(i,row).font = AR(True)
    ws.cell(i,18).font = AR(True); ws.cell(i,20).font = AR(True)
    for j in range(1,24): ws.cell(i,j).border = BORDE
fin = len(skus)+1
nota = ws.cell(fin+2,2,"Fijos USD/ud = (agente aduanal + flete interno) prorrateados entre las unidades del contenedor si fuera monoproducto; en mezcla real el prorrateo va por CBM (ver CONTENEDOR). IVA de importación no incluido en el costo: es acreditable (PARAMETROS!B11).")
nota.font = AR(sz=9, col="666666")

# ═══ CONTENEDOR (simulador del primer 40'HQ) ═══
ws = wb.create_sheet("CONTENEDOR")
ws.cell(1,1,"SIMULADOR · PRIMER CONTENEDOR 40'HQ (edita cantidades en azul)").font = AR(True,12)
cab = ["SKU","Producto","Cantidad","CBM total","Inversión aterrizada MXN","Venta mayoreo MXN","Utilidad MXN"]
anchos = [10,34,10,11,20,18,16]
for j,(t,a) in enumerate(zip(cab,anchos),1):
    c = ws.cell(3,j,t); c.font = HDR; c.fill = GRISF
    ws.column_dimensions[get_column_letter(j)].width = a
# mezcla sugerida: los 4 paquetes tipo del kit x5 + relleno de sillas
mix = {"EQP-102":200,"EQP-105":100,"EQP-101":60,"EQP-103":100,"EQP-201":60,"EQP-202":40,"EQP-301":40,"EQP-302":40,
       "EQP-404":10,"EQP-401":8,"EQP-601":6,"EQP-802":30,"EQP-501":40,"EQP-701":10,"EQP-901":4,"EQP-951":10,"EQP-960":16}
i = 4
for sku, qty in mix.items():
    ws.cell(i,1,sku).font = AR(True)
    ws.cell(i,2,f'=INDEX(CATALOGO!B:B,MATCH(A{i},CATALOGO!A:A,0))').font = AR()
    c = ws.cell(i,3,qty); c.font = AZUL; c.fill = AMAR; c.number_format='#,##0'
    ws.cell(i,4,f'=C{i}*INDEX(CATALOGO!E:E,MATCH(A{i},CATALOGO!A:A,0))').number_format='0.0'
    ws.cell(i,5,f'=C{i}*INDEX(CATALOGO!N:N,MATCH(A{i},CATALOGO!A:A,0))').number_format=MXN
    ws.cell(i,6,f'=C{i}*INDEX(CATALOGO!O:O,MATCH(A{i},CATALOGO!A:A,0))').number_format=MXN
    ws.cell(i,7,f'=F{i}-E{i}').number_format=MXN
    for j in [4,5,6]: ws.cell(i,j).font=AR()
    ws.cell(i,7).font=AR(True)
    i += 1
t = i+1
ws.cell(t,2,"TOTAL").font = AR(True,11)
ws.cell(t,3,f"=SUM(C4:C{i-1})").number_format='#,##0'
ws.cell(t,4,f"=SUM(D4:D{i-1})").number_format='0.0'
ws.cell(t,5,f"=SUM(E4:E{i-1})").number_format=MXN
ws.cell(t,6,f"=SUM(F4:F{i-1})").number_format=MXN
ws.cell(t,7,f"=SUM(G4:G{i-1})").number_format=MXN
for j in range(3,8): ws.cell(t,j).font=AR(True,11)
ws.cell(t+1,2,"CBM disponible (PARAMETROS!B10)").font=AR(sz=9,col="666666")
ws.cell(t+1,4,f"={P_(10)}").number_format='0.0'; ws.cell(t+1,4).font=AR(sz=9,col="666666")
ws.cell(t+2,2,"¿Cabe? (CBM total ≤ disponible)").font=AR(True)
ws.cell(t+2,4,f'=IF(D{t}<=D{t+1},"SÍ CABE","NO CABE — quita piezas")').font=AR(True, col="B8562E")
ws.cell(t+4,2,"IVA de importación estimado (flujo acreditable, no costo):").font=AR(sz=9,col="666666")
ws.cell(t+4,5,f"=E{t}*{P_(11)}").number_format=MXN; ws.cell(t+4,5).font=AR(sz=9,col="666666")

# ═══ COMPETENCIA ═══
ws = wb.create_sheet("COMPETENCIA")
ws.cell(1,1,"REFERENCIA: mobiliarioenmexico.com.mx — 465 productos (descarga 20-ago-2026, datos crudos en catalogo-referencia.json)").font=AR(True,11)
from collections import defaultdict
por_tipo = defaultdict(list)
for p in comp:
    precios = [float(v["price"]) for v in p["variants"] if v.get("price")]
    if precios: por_tipo[p["product_type"] or "(sin tipo)"].append((min(precios),max(precios)))
cab = ["Categoría (suya)","# productos","Precio mín MXN","Precio máx MXN"]
for j,t in enumerate(cab,1):
    c=ws.cell(3,j,t); c.font=HDR; c.fill=GRISF
    ws.column_dimensions[get_column_letter(j)].width=[34,12,16,16][j-1]
for i,(t,items) in enumerate(sorted(por_tipo.items(), key=lambda x:-len(x[1])),4):
    ws.cell(i,1,t).font=AR()
    ws.cell(i,2,len(items)).font=AR()
    ws.cell(i,3,min(x[0] for x in items)).number_format=MXN; ws.cell(i,3).font=AR()
    ws.cell(i,4,max(x[1] for x in items)).number_format=MXN; ws.cell(i,4).font=AR()

wb.save("/Users/andremacouzet/claude/equipa/investigacion/EQUIPA-CATALOGO-MAESTRO.xlsx")
print("Excel escrito:", len(skus), "SKUs")
